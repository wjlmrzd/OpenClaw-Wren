# -*- coding: utf-8 -*-
"""
Excel Data Processor - Python backend using openpyxl
"""
import sys
import os
import json
import argparse
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def read_excel(input_path, sheet=0, cell=None, range_=None):
    """Read Excel file and return data."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]
    
    if cell:
        return ws[cell].value
    
    if range_:
        result = []
        for row in ws[range_]:
            result.append([cell.value for cell in row])
        return result
    
    # Return all data
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    return data


def write_excel(output_path, data, headers=None, sheet_name="Sheet1"):
    """Write data to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if headers:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    return output_path


def format_cell(input_path, output_path, cell_ref, bold=False, italic=False,
                font_size=12, font_color=None, bg_color=None, align="left"):
    """Format a specific cell."""
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    cell = ws[cell_ref]
    
    if bold:
        cell.font = Font(bold=True)
    if italic:
        cell.font = Font(italic=True)
    if font_size:
        cell.font = Font(size=font_size)
    if font_color:
        cell.font = Font(color=font_color)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    
    align_map = {"left": "left", "center": "center", "right": "right"}
    cell.alignment = Alignment(horizontal=align_map.get(align, "left"))
    
    wb.save(output_path)
    return output_path


def create_chart(input_path, output_path, chart_type="Bar", title="Chart",
                 data_range=None, chart_position="E2"):
    """Create a chart in the Excel file."""
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    # Find data range if not specified
    if not data_range:
        max_row = ws.max_row
        max_col = ws.max_column
        data_range = f"A1:{get_column_letter(max_col)}{max_row}"
    
    # Parse range
    if ":" in data_range:
        start, end = data_range.split(":")
        # Get dimensions
        data_rows = []
        for row in ws[start:end]:
            data_rows.append([cell.value for cell in row])
        
        # Count rows with data
        data_row_count = len(data_rows)
        
        # Create chart
        if chart_type == "Bar":
            chart = BarChart()
            chart.type = "col"
        elif chart_type == "Line":
            chart = LineChart()
        elif chart_type == "Pie":
            chart = PieChart()
        else:
            chart = BarChart()
        
        chart.title = title
        chart.style = 10
        
        # Add data to chart
        data = Reference(ws, min_col=2, min_row=1, max_row=data_row_count, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_row_count)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, chart_position)
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Excel Data Processor")
    parser.add_argument("action", choices=["read", "write", "format", "chart"])
    parser.add_argument("--input", "-i")
    parser.add_argument("--output", "-o")
    parser.add_argument("--sheet", default=0, type=int)
    parser.add_argument("--cell")
    parser.add_argument("--range")
    parser.add_argument("--headers", nargs="+")
    parser.add_argument("--data", nargs="+")
    parser.add_argument("--cell-ref")
    parser.add_argument("--bold", action="store_true")
    parser.add_argument("--italic", action="store_true")
    parser.add_argument("--font-size", type=int, default=12)
    parser.add_argument("--font-color")
    parser.add_argument("--bg-color")
    parser.add_argument("--align", default="left")
    parser.add_argument("--chart-type", default="Bar")
    parser.add_argument("--chart-title", default="Chart")
    parser.add_argument("--chart-range")
    parser.add_argument("--chart-position", default="E2")
    parser.add_argument("--sheet-name", default="Sheet1")
    
    args = parser.parse_args()
    
    if args.action == "read":
        result = read_excel(args.input, args.sheet, args.cell, args.range)
        if isinstance(result, list):
            print(json.dumps(result, ensure_ascii=False, indent=2))
        else:
            print(result)
    
    elif args.action == "write":
        data = []
        if args.data:
            # Parse data pairs: col1:val1 col2:val2 ...
            for item in args.data:
                if ":" in item:
                    row_data = item.split(":")
                    data.append(row_data)
                else:
                    data.append([item])
        result = write_excel(args.output, data, args.headers, args.sheet_name)
        print(f"Written to: {result}")
    
    elif args.action == "format":
        result = format_cell(args.input, args.output, args.cell_ref,
                            args.bold, args.italic, args.font_size,
                            args.font_color, args.bg_color, args.align)
        print(f"Formatted: {result}")
    
    elif args.action == "chart":
        result = create_chart(args.input, args.output, args.chart_type,
                             args.chart_title, args.chart_range, args.chart_position)
        print(f"Chart created: {result}")


if __name__ == "__main__":
    main()
